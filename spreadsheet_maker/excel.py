from openpyxl import Workbook, load_workbook
from openpyxl.drawing.image import Image
import os
import string
import pandas as pd

__all__ = ['make_template', 'populate_template']

def make_template(fn, sheets=None, overwrite=False):
    # suffix file name with _template. this helps it to be associated programatically, and means both files can be opened in MS excel for comparison.
    
    os.makedirs(os.path.dirname('spreadsheets/excel_templates/'), exist_ok=True)
    
    name, ext = os.path.splitext(fn)
    template_fn = "{name}{suffix}{ext}".format(name=name, suffix='_template', ext=ext)

    fn = os.path.join('spreadsheets/excel_templates/', template_fn)
    wb = Workbook()
    if sheets:
        if type(sheets) == str:
            sheets = [sheets]
        ws1 = wb.active
        ws1.title = 'Contents'
        img = Image('dcms_logo.jpg')
        ws1.add_image(img, 'A1')
        ws1['B20'] = 'Contents'

        ws1.sheet_view.showGridLines = False
        for sh in sheets:
            temp_ws = wb.create_sheet(title=sh)
            temp_ws.sheet_view.showGridLines = False

        for i, sh in enumerate(sheets):
            myrow = str(21 + i)
            # ws1['B' + myrow] = '=HYPERLINK("#' + sh + '!A1","' + sh + '")'
            ws1['B' + myrow] = sh

    if os.path.exists(fn) and not overwrite:
        print('file already exists')
        return
    else:
        wb.save(filename = fn)


def populate_template(fn, cell='A6', tables={}):
    os.makedirs(os.path.dirname('spreadsheets/outputs/'), exist_ok=True)
    name, ext = os.path.splitext(fn)
    template_fn = "{name}{suffix}{ext}".format(name=name, suffix='_template', ext=ext)

    s_row = int(cell[1])
    s_col = string.ascii_uppercase.index(cell[0]) + 1

    wb = load_workbook(filename = os.path.join('spreadsheets/excel_templates/', template_fn))

    for k in tables:
        ws = wb[k]
        if isinstance(tables[k], pd.DataFrame):
            df_dict = tables[k].reset_index().to_dict('list')
            for col_idx, col in enumerate(df_dict):
                ws.cell(row=s_row, column=s_col + col_idx, value=col)
                for row_idx, row in enumerate(df_dict[col]):
                    ws.cell(row=s_row + row_idx + 1, column=s_col + col_idx, value=row)
    wb.save(filename = os.path.join('spreadsheets/outputs/', fn))


if __name__ == '__main__':
    os.chdir('publications/nov_2016')
    make_template(
        fn = 'GVA_sector_tables_template.xlsx',
        sheets=[
            "1.1 - GVA current (£bn)",
            "1.1a - GVA current (2010=100)",
            "2.1 - GVA CVM (£bn)",
            "2.1a - GVA CVM (2010=100)"],
        overwrite=True)
    make_template(
        fn = 'GVA_subsector_tables_template.xlsx',
        sheets=[
            "1 - Creative Industries-current",
            "2 - Digital Sector-current",
            "3 - Cultural Sector-current",
            "4 - Computer Games-current",
            "5 - Creative Industries-CVM",
            "6 - Digital Sector-CVM",
            "7 - Cultural Sector-CVM",],
        overwrite=True)
    import pandas as pd
    df=pd.DataFrame({'hi': [5,6], 'hello': [6,7], 'chunky': ['oops', 'babes']})
    populate_template(
        fn = 'GVA_sector_tables.xlsx',
        tables={
            "1.1 - GVA current (£bn)": df,
            "1.1a - GVA current (2010=100)": 'hi',
            "2.1 - GVA CVM (£bn)": 'lo',
            "2.1a - GVA CVM (2010=100)": 'sho',
        }
    )
# wb = load_workbook(filename = 'empty_book.xlsx')
# sheet_ranges = wb['range names']
# print(sheet_ranges['D18'].value)
